#!/usr/bin/env python3
"""
Parse 'Data prep for AI tourism.xlsx' → Firebase-ready JSON files.

Output:
  data/protected_areas.json  — list of ProtectedArea documents
  data/wildlife.json         — list of Wildlife documents
  data/sightings.json        — list of WildlifeSighting junction records

Usage:
  python3 scripts/parse_excel.py
"""

import json
import re
import ast
from pathlib import Path
import openpyxl

EXCEL_PATH = Path(__file__).parent.parent.parent / "tuarai-frontend" / "Data prep for AI tourism.xlsx"
OUT_DIR = Path(__file__).parent.parent / "data"
OUT_DIR.mkdir(exist_ok=True)

# ── Mappings ────────────────────────────────────────────────────────────────

VERTEBRATE_MAP = {
    "สัตว์เลี้ยงลูกด้วยนม": "mammal",
    "นก": "bird",
    "สัตว์เลื้อยคลาน": "reptile",
    "สัตว์สะเทินน้ำสะเทินบก": "amphibian",
}

DANGER_MAP = {
    "ต่ำ": "low",
    "ต่ำ-กลาง": "low-medium",
    "กลาง": "medium",
    "กลาง-สูง": "medium-high",
    "สูง": "high",
    "สูงมาก": "very-high",
}

# ── Helpers ──────────────────────────────────────────────────────────────────

def parse_location_json(raw) -> list[dict]:
    if not raw:
        return []
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        pass
    try:
        return ast.literal_eval(str(raw))
    except Exception:
        return []


def slugify(name: str) -> str:
    """Create a clean ID from English park name."""
    name = name.lower().strip()
    name = re.sub(r"[^a-z0-9]+", "-", name)
    return name.strip("-")


def park_id_from_name(en_name: str, park_map: dict[str, str]) -> str | None:
    """Look up a park ID by matching English name (case-insensitive, partial)."""
    en_name_lower = en_name.lower()
    for key, pid in park_map.items():
        if en_name_lower in key.lower() or key.lower() in en_name_lower:
            return pid
    return None


# ── Parse Sheet 2: Protected Areas ──────────────────────────────────────────

def parse_protected_areas(ws) -> tuple[list[dict], dict[str, str]]:
    """Returns (list_of_docs, en_name_to_id_map)."""
    docs = []
    en_name_to_id: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        (raw_id, th_name, en_name, province, gmap_url, open_close, warning, danger_raw) = (
            row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
        )

        doc_id = str(raw_id).strip()
        danger_th = str(danger_raw).strip() if danger_raw else ""
        danger_level = DANGER_MAP.get(danger_th, "medium")

        doc = {
            "id": doc_id,
            "thName": str(th_name).strip() if th_name else "",
            "enName": str(en_name).strip() if en_name else "",
            "province": str(province).strip() if province else "",
            "googleMapUrl": str(gmap_url).strip() if gmap_url else "",
            "openCloseInfo": str(open_close).strip() if open_close else "",
            "warningData": str(warning).strip() if warning else "",
            "dangerLevel": danger_level,
            "dangerLevelTh": danger_th,
            "wildlifeIds": [],  # filled in next pass
        }
        docs.append(doc)
        if en_name:
            en_name_to_id[str(en_name).strip()] = doc_id

    return docs, en_name_to_id


# ── Parse Sheet 1: Wildlife ──────────────────────────────────────────────────

def parse_wildlife(ws, en_name_to_id: dict[str, str]) -> tuple[list[dict], list[dict]]:
    """Returns (wildlife_docs, sighting_docs)."""
    wildlife_docs = []
    sighting_docs = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row[0]:
            continue
        (th_name, sci_name, common_name, vertebrate_th, loc_raw, iucn, img_url) = (
            row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        )

        vertebrate_th_str = str(vertebrate_th).strip() if vertebrate_th else ""
        vertebrate_type = VERTEBRATE_MAP.get(vertebrate_th_str, "mammal")
        iucn_status = str(iucn).strip() if iucn else "LC"
        doc_id = f"w{i+1:03d}"  # e.g. w001, w002 ...

        locations = parse_location_json(loc_raw)
        park_ids: list[str] = []

        for loc in locations:
            pa_en = loc.get("protected_area_en", "")
            pid = park_id_from_name(pa_en, en_name_to_id) if pa_en else None

            if pid and pid not in park_ids:
                park_ids.append(pid)

            sighting_docs.append({
                "wildlifeId": doc_id,
                "protectedAreaId": pid or pa_en,
                "siteCode": loc.get("site_code", ""),
                "localityTh": loc.get("locality_th", ""),
                "localityEn": loc.get("locality_en", ""),
                "province": loc.get("province", ""),
                "xIndian75": loc.get("x_indian75"),
                "yIndian75": loc.get("y_indian75"),
            })

        wildlife_docs.append({
            "id": doc_id,
            "thName": str(th_name).strip() if th_name else "",
            "enName": str(common_name).strip() if common_name else "",
            "scientificName": str(sci_name).strip() if sci_name else "",
            "vertebrateType": vertebrate_type,
            "vertebrateTypeTh": vertebrate_th_str,
            "iucnStatus": iucn_status,
            "imageUrl": str(img_url).strip() if img_url else "",
            "parkIds": park_ids,
        })

    return wildlife_docs, sighting_docs


# ── Cross-link: fill wildlifeIds on each park ────────────────────────────────

def cross_link(parks: list[dict], wildlife: list[dict]):
    park_index = {p["id"]: p for p in parks}
    for animal in wildlife:
        for pid in animal["parkIds"]:
            if pid in park_index:
                if animal["id"] not in park_index[pid]["wildlifeIds"]:
                    park_index[pid]["wildlifeIds"].append(animal["id"])


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws_parks = wb["ข้อมูลอุทยาน"]
    ws_wildlife = wb["EN VU 147_main"]

    print("Parsing protected areas...")
    parks, en_name_to_id = parse_protected_areas(ws_parks)

    print("Parsing wildlife...")
    wildlife, sightings = parse_wildlife(ws_wildlife, en_name_to_id)

    print("Cross-linking...")
    cross_link(parks, wildlife)

    # Write outputs
    (OUT_DIR / "protected_areas.json").write_text(json.dumps(parks, ensure_ascii=False, indent=2))
    (OUT_DIR / "wildlife.json").write_text(json.dumps(wildlife, ensure_ascii=False, indent=2))
    (OUT_DIR / "sightings.json").write_text(json.dumps(sightings, ensure_ascii=False, indent=2))

    print(f"\nDone!")
    print(f"  Protected areas : {len(parks)}")
    print(f"  Wildlife        : {len(wildlife)}")
    print(f"  Sightings       : {len(sightings)}")
    print(f"\nJSON files written to: {OUT_DIR}")


if __name__ == "__main__":
    main()
